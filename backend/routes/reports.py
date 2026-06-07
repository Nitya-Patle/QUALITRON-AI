"""
Reports Routes — /api/reports
GET /pdf    download PDF report
GET /excel  download Excel workbook
"""

import io
from datetime import datetime, timedelta, timezone
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from database.db import get_db

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from collections import defaultdict

reports_bp = Blueprint("reports", __name__)


def _records(days=7):
    db    = get_db()
    operator = get_jwt_identity()
    since = datetime.now(timezone.utc) - timedelta(days=days)
    return list(db.inspections.find({"operator": operator, "timestamp": {"$gte": since}}).sort("timestamp", -1).limit(500))


# ── PDF ──────────────────────────────────────────────────────
@reports_bp.route("/pdf", methods=["GET"])
@jwt_required()
def pdf():
    days    = int(request.args.get("days", 7))
    records = _records(days)

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                              topMargin=2*cm, bottomMargin=2*cm,
                              leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    title_style = ParagraphStyle("t", fontSize=18, textColor=colors.HexColor("#00D4FF"),
                                  fontName="Helvetica-Bold", spaceAfter=4)
    story.append(Paragraph("QUALITRON AI — Quality Control Report", title_style))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Period: Last {days} days",
        styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1e2d54")))
    story.append(Spacer(1, 0.3*cm))

    total     = len(records)
    defective = sum(1 for r in records if r["status"] == "FAIL")
    passed    = total - defective
    avg_acc   = (sum(r.get("accuracy", 0) for r in records) / total) if total else 0

    kpi_data = [
        ["Metric", "Value"],
        ["Total Inspected",  str(total)],
        ["Passed",           str(passed)],
        ["Defective",        str(defective)],
        ["Defect Rate",      f"{(defective/total*100):.1f}%" if total else "N/A"],
        ["Avg AI Accuracy",  f"{avg_acc:.1f}%"],
    ]
    kpi_t = Table(kpi_data, colWidths=[7*cm, 5*cm])
    kpi_t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0f1628")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.HexColor("#00D4FF")),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#141c35")),
        ("TEXTCOLOR",  (0,1), (-1,-1), colors.white),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#1e2d54")),
        ("FONTSIZE",   (0,0), (-1,-1), 10),
    ]))
    story.append(kpi_t)
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("Inspection Records", styles["Heading2"]))
    story.append(Spacer(1, 0.2*cm))

    rows = [["ID", "Product", "Operator", "Status", "Defects", "Accuracy", "Date"]]
    for r in records[:100]:
        rows.append([
            str(r["_id"])[:12],
            r.get("product", "—"),
            r.get("operator", "—"),
            r.get("status", "—"),
            str(r.get("defect_count", 0)),
            f"{r.get('accuracy',0):.1f}%",
            r["timestamp"].strftime("%d/%m/%y %H:%M"),
        ])
    rec_t = Table(rows, repeatRows=1,
                  colWidths=[3*cm, 3.5*cm, 2.5*cm, 2*cm, 2*cm, 2.5*cm, 3*cm])
    rec_t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0f1628")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.HexColor("#00D4FF")),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("TEXTCOLOR",  (0,1), (-1,-1), colors.white),
        ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#1e2d54")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1),
         [colors.HexColor("#141c35"), colors.HexColor("#0f1628")]),
    ]))
    story.append(rec_t)
    doc.build(story)
    buf.seek(0)
    fname = f"qualitron_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fname)


# ── EXCEL ────────────────────────────────────────────────────
@reports_bp.route("/excel", methods=["GET"])
@jwt_required()
def excel():
    days    = int(request.args.get("days", 7))
    records = _records(days)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspections"

    hfill  = PatternFill("solid", fgColor="0F1628")
    hfont  = Font(bold=True, color="00D4FF", size=11)
    border = Border(
        left=Side(style="thin", color="1E2D54"), right=Side(style="thin", color="1E2D54"),
        top=Side(style="thin",  color="1E2D54"), bottom=Side(style="thin", color="1E2D54"),
    )
    headers = ["ID","Product","Operator","Station","Status","Defects","Accuracy(%)","Model","Timestamp"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill
        c.alignment = Alignment(horizontal="center"); c.border = border

    for ri, r in enumerate(records, 2):
        fc = "0A2014" if r.get("status") == "PASS" else "200A0E"
        vals = [str(r["_id"])[:16], r.get("product",""), r.get("operator",""),
                r.get("station","A"), r.get("status",""),
                r.get("defect_count",0), r.get("accuracy",0),
                r.get("model","YOLOv8x"), r["timestamp"].strftime("%d/%m/%Y %H:%M:%S")]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(color="E8EEFF", size=10)
            c.fill = PatternFill("solid", fgColor=fc)
            c.alignment = Alignment(horizontal="center"); c.border = border

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    ws2 = wb.create_sheet("Daily Summary")
    ws2.append(["Date","Total","Defective","Pass Rate %"])
    daily = defaultdict(lambda: {"total":0,"defective":0})
    for r in records:
        day = r["timestamp"].strftime("%Y-%m-%d")
        daily[day]["total"]     += 1
        daily[day]["defective"] += 1 if r.get("status") == "FAIL" else 0
    for day in sorted(daily):
        t = daily[day]["total"]; d = daily[day]["defective"]
        ws2.append([day, t, d, round((t-d)/t*100, 1) if t else 0])

    chart = BarChart()
    chart.title = "Daily Inspection Summary"
    chart.y_axis.title = "Count"
    data = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "F2")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"qualitron_data_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname)
